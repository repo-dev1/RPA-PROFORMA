"""
Excel generator for the Cuadro Comparativo (Comparative Quote Table).
Replicates the exact format of the ACONIC template.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage


# ── Style Constants ──────────────────────────────────────────────────────────

# ACONIC blue (theme color 4 approximation)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

HEADER_FONT_LARGE = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HEADER_FONT_SMALL = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
DESC_FONT = Font(name="ArialMT", size=8, bold=False)
DATA_FONT = Font(name="Calibri", size=9, bold=True)
SUMMARY_FONT = Font(name="Calibri", size=10, bold=True)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Nicaraguan Córdoba currency format
CS_FORMAT = '_-[$C$-4C0A]* #,##0.00_-;\\-[$C$-4C0A]* #,##0.00_-;_-[$C$-4C0A]* "-"??_-;_-@_-'

# Column widths matching the template
COLUMN_WIDTHS = {
    'A': 4.0,      # No
    'B': 36.0,     # Descripción
    'C': 12.5,     # Proveedor 1 (precio unitario)
    'D': 13.0,     # Proveedor 2 (precio unitario)
    'E': 12.5,     # Proveedor 3 (precio unitario)
    'F': 8.0,      # CANTIDAD
    'G': 13.5,     # Proveedor 1 (total)
    'H': 13.0,     # Proveedor 2 (total)
    'I': 14.0,     # Proveedor 3 (total)
}


def generate_comparativo(
    items_data: list,
    suppliers: list,
    title: str = "PRODUCTOS VARIOS",
    output_path: str = "output/Cuadro_Comparativo.xlsx",
    logo_path: str = None,
) -> str:
    """
    Generate the Cuadro Comparativo Excel file.

    Args:
        items_data: List of dicts with keys:
            - descripcion (str)
            - cantidad (int/float)
            - precios: dict {supplier_name: precio_unitario} (only for suppliers that quoted)
        suppliers: List of 3 supplier names, e.g. ["MASTERTEC", "CONICO", "COMTECH"]
        title: Title for the comparative table
        output_path: Where to save the Excel file
        logo_path: Path to the ACONIC logo image

    Returns:
        Absolute path of the generated file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"

    # Ensure we have exactly 3 suppliers
    while len(suppliers) < 3:
        suppliers.append(f"PROVEEDOR {len(suppliers) + 1}")
    suppliers = suppliers[:3]

    # ── Column widths ────────────────────────────────────────────────────────
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Row 1: Empty ─────────────────────────────────────────────────────────
    # (space for logo)

    # ── Rows 2-4: Title (merged B2:I4) ───────────────────────────────────────
    ws.merge_cells('B2:I4')
    title_cell = ws['B2']
    title_cell.value = f"Cuadro comparativo  - {title}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Insert logo if available
    if logo_path and os.path.exists(logo_path):
        try:
            img = XlImage(logo_path)
            img.width = 100
            img.height = 100
            ws.add_image(img, 'B2')
        except Exception as e:
            print(f"Warning: Could not add logo: {e}")

    # ── Rows 5-6: Empty ─────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    # ── Row 7: "PROVEEDORES" header (merged C7:I7) ──────────────────────────
    ws.merge_cells('C7:I7')
    prov_cell = ws['C7']
    prov_cell.value = "PROVEEDORES"
    prov_cell.font = HEADER_FONT_LARGE
    prov_cell.fill = HEADER_FILL
    prov_cell.alignment = CENTER_ALIGN
    # Apply fill to all merged cells in row 7
    for col in range(3, 10):  # C to I
        cell = ws.cell(row=7, column=col)
        cell.fill = HEADER_FILL

    # ── Row 8: Column headers ────────────────────────────────────────────────
    headers = [
        ("A8", "No", HEADER_FONT_LARGE),
        ("B8", "Descripción", HEADER_FONT_LARGE),
        ("C8", suppliers[0], HEADER_FONT_SMALL),
        ("D8", suppliers[1], HEADER_FONT_SMALL),
        ("E8", suppliers[2], HEADER_FONT_SMALL),
        ("F8", "CANTIDAD", HEADER_FONT_SMALL),
        ("G8", suppliers[0], HEADER_FONT_SMALL),
        ("H8", suppliers[1], HEADER_FONT_SMALL),
        ("I8", suppliers[2], HEADER_FONT_SMALL),
    ]
    for cell_ref, value, font in headers:
        cell = ws[cell_ref]
        cell.value = value
        cell.font = font
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    ws.row_dimensions[8].height = 30

    # ── Data rows (starting at row 9) ────────────────────────────────────────
    data_start_row = 9

    for idx, item in enumerate(items_data):
        row = data_start_row + idx
        desc = item.get('descripcion', '')
        cant = item.get('cantidad', 0)
        precios = item.get('precios', {})

        # Column A: Number
        cell_a = ws.cell(row=row, column=1, value=idx + 1)
        cell_a.font = DATA_FONT
        cell_a.alignment = CENTER_ALIGN
        cell_a.border = THIN_BORDER

        # Column B: Description
        cell_b = ws.cell(row=row, column=2, value=desc)
        cell_b.font = DESC_FONT
        cell_b.alignment = LEFT_ALIGN
        cell_b.border = THIN_BORDER

        # Column C: Supplier 1 unit price
        precio_1 = precios.get(suppliers[0], None)
        cell_c = ws.cell(row=row, column=3, value=precio_1 if precio_1 else None)
        cell_c.font = DATA_FONT
        cell_c.fill = WHITE_FILL
        cell_c.number_format = CS_FORMAT
        cell_c.alignment = CENTER_ALIGN
        cell_c.border = THIN_BORDER

        # Column D: Supplier 2 unit price
        precio_2 = precios.get(suppliers[1], None)
        cell_d = ws.cell(row=row, column=4, value=precio_2 if precio_2 else None)
        cell_d.font = DATA_FONT
        cell_d.fill = WHITE_FILL
        cell_d.number_format = CS_FORMAT
        cell_d.alignment = CENTER_ALIGN
        cell_d.border = THIN_BORDER

        # Column E: Supplier 3 unit price
        precio_3 = precios.get(suppliers[2], None)
        cell_e = ws.cell(row=row, column=5, value=precio_3 if precio_3 else None)
        cell_e.font = DATA_FONT
        cell_e.fill = WHITE_FILL
        cell_e.number_format = CS_FORMAT
        cell_e.alignment = CENTER_ALIGN
        cell_e.border = THIN_BORDER

        # Column F: Quantity
        cell_f = ws.cell(row=row, column=6, value=cant if cant else None)
        cell_f.font = DATA_FONT
        cell_f.fill = WHITE_FILL
        cell_f.alignment = CENTER_ALIGN
        cell_f.border = THIN_BORDER

        # Column G: Supplier 1 total (formula)
        col_c = get_column_letter(3)  # C
        col_f = get_column_letter(6)  # F
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = f"={col_c}{row}*{col_f}{row}"
        cell_g.font = DATA_FONT
        cell_g.fill = WHITE_FILL
        cell_g.number_format = CS_FORMAT
        cell_g.alignment = CENTER_ALIGN
        cell_g.border = THIN_BORDER

        # Column H: Supplier 2 total (formula)
        col_d = get_column_letter(4)  # D
        cell_h = ws.cell(row=row, column=8)
        cell_h.value = f"={col_d}{row}*{col_f}{row}"
        cell_h.font = DATA_FONT
        cell_h.fill = WHITE_FILL
        cell_h.number_format = CS_FORMAT
        cell_h.alignment = CENTER_ALIGN
        cell_h.border = THIN_BORDER

        # Column I: Supplier 3 total (formula)
        col_e = get_column_letter(5)  # E
        cell_i = ws.cell(row=row, column=9)
        cell_i.value = f"={col_e}{row}*{col_f}{row}"
        cell_i.font = DATA_FONT
        cell_i.fill = WHITE_FILL
        cell_i.number_format = CS_FORMAT
        cell_i.alignment = CENTER_ALIGN
        cell_i.border = THIN_BORDER

        # Set row height for wrapped text
        ws.row_dimensions[row].height = 30

    # ── Summary rows: SUBTOTAL, IVA, TOTAL ───────────────────────────────────
    data_end_row = data_start_row + len(items_data) - 1
    if len(items_data) == 0:
        data_end_row = data_start_row  # At least one row

    summary_labels = ["SUBTOTAL", "IVA", "TOTAL"]
    for s_idx, label in enumerate(summary_labels):
        row = data_end_row + 1 + s_idx

        # Label in column F
        cell_label = ws.cell(row=row, column=6, value=label)
        cell_label.font = SUMMARY_FONT
        cell_label.alignment = CENTER_ALIGN
        cell_label.border = THIN_BORDER

        for col in [7, 8, 9]:  # G, H, I
            col_letter = get_column_letter(col)
            cell = ws.cell(row=row, column=col)

            if label == "SUBTOTAL":
                cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
            elif label == "IVA":
                cell.value = f"=+{col_letter}{row - 1}*0.15"
            elif label == "TOTAL":
                cell.value = f"=+{col_letter}{row - 2}+{col_letter}{row - 1}"

            cell.font = SUMMARY_FONT
            cell.fill = WHITE_FILL
            cell.number_format = CS_FORMAT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

    # ── Print setup ──────────────────────────────────────────────────────────
    from openpyxl.worksheet.properties import PageSetupProperties

    # Enable "Fit to Page" mode (critical — without this, fitToWidth is ignored)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = auto (as many pages tall as needed)

    # Set print area to include all content
    last_data_row = data_end_row + 3  # +3 for SUBTOTAL, IVA, TOTAL
    ws.print_area = f"A1:I{last_data_row}"

    # Narrow margins for more print space
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    # Center content on page
    ws.print_options.horizontalCentered = True

    # ── Save ─────────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)

    return os.path.abspath(output_path)
