"""
Excel generator for the Orden de Compra (Purchase Order).
Replicates the exact format of the ACONIC OC template.
"""

import os
from datetime import date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage


# ── Style Constants ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
COMPANY_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=9)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CS_FORMAT = '_-[$C$-4C0A]* #,##0.00_-;\\-[$C$-4C0A]* #,##0.00_-;_-[$C$-4C0A]* "-"??_-;_-@_-'


def generate_orden_compra(
    items_data: list,
    supplier: str,
    solicitante: str = "David Garcia",
    area: str = "Informatica",
    tipo_compra: str = "Equipos de computo",
    autoriza: str = "Ivette Medina",
    output_path: str = "output/Orden_de_Compra.xlsx",
    logo_path: str = None,
) -> str:
    """
    Generate the Orden de Compra Excel file by loading the existing ACONIC template.
    """
    import glob
    import shutil
    
    # Locate the template file
    template_files = glob.glob(os.path.join("templates", "OC*.xlsx"))
    if not template_files:
        raise FileNotFoundError("No se encontró ninguna plantilla que empiece con 'OC' en la carpeta templates.")
    
    template_path = template_files[0]
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Replace the text fields
    ws['B10'].value = supplier
    ws['A12'].value = f" {supplier}"
    
    ws['B15'].value = date.today().strftime('%d/%m/%Y')
    ws['C15'].value = solicitante
    ws['D15'].value = tipo_compra
    ws['E15'].value = area
    ws['F15'].value = autoriza
    ws['G13'].value = solicitante
    
    # Clear existing item rows (22 to 29) to be safe
    for row in range(22, 30):
        ws.cell(row=row, column=1).value = None  # A
        ws.cell(row=row, column=2).value = None  # B
        ws.cell(row=row, column=5).value = None  # E
        ws.cell(row=row, column=6).value = None  # F
        ws.cell(row=row, column=7).value = None  # G
    
    # Insert new items
    for idx, item in enumerate(items_data):
        if idx >= 8:  # Template supports max 8 items visually
            break
            
        row = 22 + idx
        
        ws.cell(row=row, column=2).value = item.get('descripcion', '')
        ws.cell(row=row, column=5).value = item.get('cantidad', 0)
        ws.cell(row=row, column=6).value = item.get('precio_unitario', 0)
        
        # Add the total formula
        ws.cell(row=row, column=7).value = f"=F{row}*E{row}"
        
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    
    return os.path.abspath(output_path)

